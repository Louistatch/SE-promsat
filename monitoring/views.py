from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponseForbidden
from .models import Realisation, Indicateur, Periode, Rapport, AlerteQualite
from .utils import calculer_synthese_nationale, generer_rapport_qualite

@login_required
def saisie_realisation_view(request):
    if request.method == 'POST':
        indicateur_id = request.POST.get('indicateur')
        periode_id = request.POST.get('periode')
        valeur = request.POST.get('valeur')
        hommes = request.POST.get('hommes', 0)
        femmes = request.POST.get('femmes', 0)
        commentaire = request.POST.get('commentaire', '')
        
        indicateur = get_object_or_404(Indicateur, id=indicateur_id)
        periode = get_object_or_404(Periode, id=periode_id)
        
        # Déterminer la région
        if request.user.has_region_access():
            region = request.user.region
        else:
            region = request.POST.get('region')
        
        # Créer ou mettre à jour la réalisation
        realisation, created = Realisation.objects.update_or_create(
            indicateur=indicateur,
            periode=periode,
            region=region,
            defaults={
                'valeur_realisee': valeur,
                'hommes': hommes,
                'femmes': femmes,
                'commentaire': commentaire,
                'saisi_par': request.user,
            }
        )
        
        # Gérer le fichier justificatif
        if request.FILES.get('fichier'):
            realisation.fichier_justificatif = request.FILES['fichier']
            realisation.save()
        
        # Vérifier la qualité et créer des alertes
        from .utils import verifier_qualite_realisation
        verifier_qualite_realisation(realisation)
        
        messages.success(request, 'Réalisation enregistrée avec succès!')
        return redirect('monitoring:liste_realisations')
    
    # GET request
    indicateurs = Indicateur.objects.filter(actif=True).select_related('sous_composante')
    periodes = Periode.objects.filter(cloture=False)
    
    context = {
        'indicateurs': indicateurs,
        'periodes': periodes,
    }
    
    return render(request, 'monitoring/saisie_realisation.html', context)

@login_required
def liste_realisations_view(request):
    if request.user.has_region_access():
        realisations = Realisation.objects.filter(region=request.user.region)
    else:
        realisations = Realisation.objects.all()
    
    realisations = realisations.select_related(
        'indicateur__sous_composante',
        'periode',
        'saisi_par'
    ).order_by('-date_saisie')
    
    # Filtres
    periode_id = request.GET.get('periode')
    if periode_id:
        realisations = realisations.filter(periode_id=periode_id)
    
    valide = request.GET.get('valide')
    if valide == 'oui':
        realisations = realisations.filter(valide=True)
    elif valide == 'non':
        realisations = realisations.filter(valide=False)
    
    periodes = Periode.objects.all()
    
    context = {
        'realisations': realisations,
        'periodes': periodes,
    }
    
    return render(request, 'monitoring/liste_realisations.html', context)

@login_required
def modifier_realisation_view(request, pk):
    realisation = get_object_or_404(Realisation, pk=pk)
    
    # Vérifier les permissions
    if request.user.has_region_access() and realisation.region != request.user.region:
        return HttpResponseForbidden("Vous n'avez pas accès à cette réalisation.")
    
    if request.method == 'POST':
        realisation.valeur_realisee = request.POST.get('valeur')
        realisation.commentaire = request.POST.get('commentaire', '')
        
        if request.FILES.get('fichier'):
            realisation.fichier_justificatif = request.FILES['fichier']
        
        realisation.save()
        messages.success(request, 'Réalisation modifiée avec succès!')
        return redirect('monitoring:liste_realisations')
    
    context = {
        'realisation': realisation,
    }
    
    return render(request, 'monitoring/modifier_realisation.html', context)

@login_required
def valider_realisation_view(request, pk):
    if not request.user.has_full_access():
        return HttpResponseForbidden("Vous n'avez pas les permissions pour valider.")
    
    realisation = get_object_or_404(Realisation, pk=pk)
    
    if request.method == 'POST':
        realisation.valide = True
        realisation.valide_par = request.user
        realisation.save()
        messages.success(request, 'Réalisation validée avec succès!')
    
    return redirect('monitoring:liste_realisations')

@login_required
def liste_rapports_view(request):
    if request.user.has_region_access():
        rapports = Rapport.objects.filter(region=request.user.region)
    else:
        rapports = Rapport.objects.all()
    
    rapports = rapports.select_related('periode', 'auteur').order_by('-date_creation')
    
    context = {
        'rapports': rapports,
    }
    
    return render(request, 'monitoring/liste_rapports.html', context)

@login_required
def detail_rapport_view(request, pk):
    rapport = get_object_or_404(Rapport, pk=pk)
    
    # Vérifier les permissions
    if request.user.has_region_access() and rapport.region != request.user.region:
        return HttpResponseForbidden("Vous n'avez pas accès à ce rapport.")
    
    context = {
        'rapport': rapport,
    }
    
    return render(request, 'monitoring/detail_rapport.html', context)


@login_required
def synthese_nationale_view(request):
    """Vue de la synthèse nationale agrégée"""
    if not request.user.has_full_access():
        return HttpResponseForbidden("Accès réservé aux coordonnateurs et évaluateurs.")
    
    # Filtres
    periode_id = request.GET.get('periode')
    periode = None
    if periode_id:
        periode = get_object_or_404(Periode, id=periode_id)
    else:
        periode = Periode.objects.filter(cloture=False).first()
    
    # Récupérer tous les indicateurs actifs
    indicateurs = Indicateur.objects.filter(actif=True).select_related('sous_composante__composante')
    
    # Calculer la synthèse pour chaque indicateur
    syntheses = []
    for indicateur in indicateurs:
        synthese = calculer_synthese_nationale(indicateur, periode)
        synthese['indicateur'] = indicateur
        syntheses.append(synthese)
    
    # Statistiques globales
    total_indicateurs = len(syntheses)
    indicateurs_atteints = sum(1 for s in syntheses if s['pourcentage_atteinte'] >= 100)
    indicateurs_en_cours = sum(1 for s in syntheses if 50 <= s['pourcentage_atteinte'] < 100)
    indicateurs_retard = sum(1 for s in syntheses if s['pourcentage_atteinte'] < 50)
    
    # Rapport qualité
    rapport_qualite = generer_rapport_qualite(periode=periode)
    
    periodes = Periode.objects.all()
    
    context = {
        'syntheses': syntheses,
        'periode': periode,
        'periodes': periodes,
        'stats': {
            'total': total_indicateurs,
            'atteints': indicateurs_atteints,
            'en_cours': indicateurs_en_cours,
            'retard': indicateurs_retard,
        },
        'rapport_qualite': rapport_qualite,
    }
    
    return render(request, 'monitoring/synthese_nationale.html', context)

@login_required
def controle_qualite_view(request):
    """Vue du contrôle qualité"""
    if not request.user.has_full_access():
        return HttpResponseForbidden("Accès réservé aux coordonnateurs et évaluateurs.")
    
    # Filtres
    periode_id = request.GET.get('periode')
    region = request.GET.get('region')
    type_alerte = request.GET.get('type')
    
    # Alertes non résolues
    alertes = AlerteQualite.objects.filter(resolue=False).select_related(
        'realisation__indicateur',
        'realisation__periode'
    )
    
    if periode_id:
        alertes = alertes.filter(realisation__periode_id=periode_id)
    
    if region:
        alertes = alertes.filter(realisation__region=region)
    
    if type_alerte:
        alertes = alertes.filter(type_alerte=type_alerte)
    
    # Rapport qualité
    rapport = generer_rapport_qualite(
        periode=Periode.objects.get(id=periode_id) if periode_id else None,
        region=region
    )
    
    periodes = Periode.objects.all()
    regions = ['MARITIME', 'PLATEAUX', 'CENTRALE', 'KARA', 'SAVANES']
    
    context = {
        'alertes': alertes,
        'rapport': rapport,
        'periodes': periodes,
        'regions': regions,
        'type_alertes': AlerteQualite.TYPE_CHOICES,
    }
    
    return render(request, 'monitoring/controle_qualite.html', context)

@login_required
def resoudre_alerte_view(request, pk):
    """Résoudre une alerte qualité"""
    if not request.user.has_full_access():
        return HttpResponseForbidden("Accès réservé aux coordonnateurs et évaluateurs.")
    
    alerte = get_object_or_404(AlerteQualite, pk=pk)
    
    if request.method == 'POST':
        commentaire = request.POST.get('commentaire', '')
        alerte.resoudre(request.user, commentaire)
        messages.success(request, 'Alerte résolue avec succès!')
        return redirect('monitoring:controle_qualite')
    
    return redirect('monitoring:controle_qualite')


@login_required
def export_excel_view(request):
    """
    Export des données en format Excel
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from django.db.models import Sum
    from datetime import datetime
    
    # Vérifier les permissions
    if not request.user.has_full_access():
        return HttpResponseForbidden("Accès refusé")
    
    # Créer le workbook
    wb = Workbook()
    
    # === FEUILLE 1: SYNTHÈSE NATIONALE ===
    ws_synthese = wb.active
    ws_synthese.title = "Synthese-Nationale"
    
    # En-têtes
    headers = [
        'Code', 'Indicateur', 'Unité', 'Cible', 'T1 2026', 'T2 2026', 'T3 2026', 'T4 2026',
        'Total Réalisé', '% Atteinte', 'Écart', 'Hommes', 'Femmes', '% Femmes'
    ]
    
    # Style des en-têtes
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_synthese.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Données
    indicateurs = Indicateur.objects.filter(actif=True).order_by('code')
    periodes = Periode.objects.all().order_by('-annee', '-trimestre')[:4]
    
    row_num = 2
    for indicateur in indicateurs:
        # Calculer les totaux nationaux
        total_realise = 0
        total_hommes = 0
        total_femmes = 0
        
        col_num = 1
        ws_synthese.cell(row=row_num, column=col_num).value = indicateur.code
        col_num += 1
        ws_synthese.cell(row=row_num, column=col_num).value = indicateur.libelle
        col_num += 1
        ws_synthese.cell(row=row_num, column=col_num).value = indicateur.unite_mesure
        col_num += 1
        ws_synthese.cell(row=row_num, column=col_num).value = float(indicateur.cible_finale or 0)
        col_num += 1
        
        # Réalisations par période
        for periode in periodes:
            synthese = calculer_synthese_nationale(indicateur, periode)
            ws_synthese.cell(row=row_num, column=col_num).value = float(synthese['total_realise'])
            total_realise += float(synthese['total_realise'])
            total_hommes += float(synthese['total_hommes'])
            total_femmes += float(synthese['total_femmes'])
            col_num += 1
        
        # Total réalisé
        ws_synthese.cell(row=row_num, column=col_num).value = float(total_realise)
        col_num += 1
        
        # % Atteinte
        pct_atteinte = (float(total_realise) / float(indicateur.cible_finale) * 100) if indicateur.cible_finale else 0
        ws_synthese.cell(row=row_num, column=col_num).value = round(pct_atteinte, 1)
        col_num += 1
        
        # Écart
        ecart = float(indicateur.cible_finale or 0) - float(total_realise)
        ws_synthese.cell(row=row_num, column=col_num).value = float(ecart)
        col_num += 1
        
        # Hommes
        ws_synthese.cell(row=row_num, column=col_num).value = float(total_hommes)
        col_num += 1
        
        # Femmes
        ws_synthese.cell(row=row_num, column=col_num).value = float(total_femmes)
        col_num += 1
        
        # % Femmes
        pct_femmes = (total_femmes / total_realise * 100) if total_realise > 0 else 0
        ws_synthese.cell(row=row_num, column=col_num).value = round(pct_femmes, 1)
        
        row_num += 1
    
    # Ajuster la largeur des colonnes
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        ws_synthese.column_dimensions[column_letter].width = 15
    
    ws_synthese.column_dimensions['B'].width = 50  # Colonne Indicateur plus large
    
    # === FEUILLE 2: SUIVI PAR RÉGION ===
    regions = ['MARITIME', 'PLATEAUX', 'CENTRALE', 'KARA', 'SAVANES']
    
    for region in regions:
        ws_region = wb.create_sheet(title=f"Suivi-{region}")
        
        # En-têtes
        for col_num, header in enumerate(headers, 1):
            cell = ws_region.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Données
        row_num = 2
        for indicateur in indicateurs:
            total_realise = 0
            total_hommes = 0
            total_femmes = 0
            
            col_num = 1
            ws_region.cell(row=row_num, column=col_num).value = indicateur.code
            col_num += 1
            ws_region.cell(row=row_num, column=col_num).value = indicateur.libelle
            col_num += 1
            ws_region.cell(row=row_num, column=col_num).value = indicateur.unite_mesure
            col_num += 1
            ws_region.cell(row=row_num, column=col_num).value = float(indicateur.cible_finale or 0)
            col_num += 1
            
            # Réalisations par période
            for periode in periodes:
                realisation = Realisation.objects.filter(
                    indicateur=indicateur,
                    periode=periode,
                    region=region
                ).first()
                
                valeur = float(realisation.valeur_realisee) if realisation else 0
                ws_region.cell(row=row_num, column=col_num).value = valeur
                total_realise += valeur
                
                if realisation:
                    total_hommes += float(realisation.hommes)
                    total_femmes += float(realisation.femmes)
                
                col_num += 1
            
            # Total réalisé
            ws_region.cell(row=row_num, column=col_num).value = float(total_realise)
            col_num += 1
            
            # % Atteinte
            pct_atteinte = (float(total_realise) / float(indicateur.cible_finale) * 100) if indicateur.cible_finale else 0
            ws_region.cell(row=row_num, column=col_num).value = round(pct_atteinte, 1)
            col_num += 1
            
            # Écart
            ecart = float(indicateur.cible_finale or 0) - float(total_realise)
            ws_region.cell(row=row_num, column=col_num).value = float(ecart)
            col_num += 1
            
            # Hommes
            ws_region.cell(row=row_num, column=col_num).value = float(total_hommes)
            col_num += 1
            
            # Femmes
            ws_region.cell(row=row_num, column=col_num).value = float(total_femmes)
            col_num += 1
            
            # % Femmes
            pct_femmes = (total_femmes / total_realise * 100) if total_realise > 0 else 0
            ws_region.cell(row=row_num, column=col_num).value = round(pct_femmes, 1)
            
            row_num += 1
        
        # Ajuster la largeur des colonnes
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            ws_region.column_dimensions[column_letter].width = 15
        
        ws_region.column_dimensions['B'].width = 50
    
    # === FEUILLE 3: CONTRÔLE QUALITÉ ===
    ws_qualite = wb.create_sheet(title="Controle-Qualite")
    
    # En-têtes
    qualite_headers = ['Région', 'Indicateur', 'Période', 'Type Alerte', 'Sévérité', 'Message', 'Date']
    
    for col_num, header in enumerate(qualite_headers, 1):
        cell = ws_qualite.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Données
    alertes = AlerteQualite.objects.filter(resolue=False).select_related(
        'realisation__indicateur', 'realisation__periode'
    ).order_by('-date_detection')
    
    row_num = 2
    for alerte in alertes:
        ws_qualite.cell(row=row_num, column=1).value = alerte.realisation.region
        ws_qualite.cell(row=row_num, column=2).value = alerte.realisation.indicateur.libelle
        ws_qualite.cell(row=row_num, column=3).value = str(alerte.realisation.periode)
        ws_qualite.cell(row=row_num, column=4).value = alerte.get_type_alerte_display()
        ws_qualite.cell(row=row_num, column=5).value = alerte.get_severite_display()
        ws_qualite.cell(row=row_num, column=6).value = alerte.message
        ws_qualite.cell(row=row_num, column=7).value = alerte.date_detection.strftime('%d/%m/%Y %H:%M')
        row_num += 1
    
    # Ajuster la largeur des colonnes
    for col_num in range(1, len(qualite_headers) + 1):
        column_letter = get_column_letter(col_num)
        ws_qualite.column_dimensions[column_letter].width = 20
    
    ws_qualite.column_dimensions['F'].width = 50  # Message plus large
    
    # Préparer la réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'ProSMAT_Export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
def export_pdf_view(request):
    """
    Export d'un rapport PDF professionnel
    """
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.platypus import Image as RLImage
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from django.http import HttpResponse
    from django.db.models import Sum
    from datetime import datetime
    import io
    
    # Vérifier les permissions
    if not request.user.has_full_access():
        return HttpResponseForbidden("Accès refusé")
    
    # Créer le buffer
    buffer = io.BytesIO()
    
    # Créer le document PDF
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1.5*cm,
        leftMargin=1.5*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#366092'),
        spaceAfter=30,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#366092'),
        spaceAfter=12,
        spaceBefore=12,
        fontName='Helvetica-Bold'
    )
    
    # Contenu du PDF
    story = []
    
    # Titre
    title = Paragraph("RAPPORT DE SUIVI-ÉVALUATION<br/>ProSMAT - TOGO", title_style)
    story.append(title)
    
    # Date
    date_text = Paragraph(
        f"<b>Date:</b> {datetime.now().strftime('%d/%m/%Y')}<br/>"
        f"<b>Période:</b> Année 2026",
        styles['Normal']
    )
    story.append(date_text)
    story.append(Spacer(1, 0.5*cm))
    
    # === SECTION 1: SYNTHÈSE EXÉCUTIVE ===
    story.append(Paragraph("1. SYNTHÈSE EXÉCUTIVE", heading_style))
    
    # Calculer les KPI
    indicateurs = Indicateur.objects.filter(actif=True)
    total_indicateurs = indicateurs.count()
    
    realisations = Realisation.objects.filter(valide=True)
    total_realisations = realisations.count()
    
    # Performance globale
    total_pct = 0
    count_ind = 0
    
    for ind in indicateurs:
        if ind.cible_finale and ind.cible_finale > 0:
            total_realise = realisations.filter(indicateur=ind).aggregate(
                Sum('valeur_realisee')
            )['valeur_realisee__sum'] or 0
            pct = (total_realise / ind.cible_finale * 100)
            total_pct += pct
            count_ind += 1
    
    performance_globale = (total_pct / count_ind) if count_ind > 0 else 0
    
    # Tableau KPI
    kpi_data = [
        ['Indicateur', 'Valeur'],
        ['Nombre d\'indicateurs', str(total_indicateurs)],
        ['Nombre de réalisations', str(total_realisations)],
        ['Performance globale', f'{performance_globale:.1f}%'],
    ]
    
    kpi_table = Table(kpi_data, colWidths=[12*cm, 8*cm])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    story.append(kpi_table)
    story.append(Spacer(1, 1*cm))
    
    # === SECTION 2: PERFORMANCE PAR RÉGION ===
    story.append(Paragraph("2. PERFORMANCE PAR RÉGION", heading_style))
    
    regions = ['MARITIME', 'PLATEAUX', 'CENTRALE', 'KARA', 'SAVANES']
    region_data = [['Région', 'Nb Réalisations', 'Performance (%)']]
    
    for region in regions:
        realisations_region = realisations.filter(region=region)
        nb_real = realisations_region.count()
        
        total_pct_region = 0
        count_ind_region = 0
        
        for ind in indicateurs:
            if ind.cible_finale and ind.cible_finale > 0:
                total_realise = realisations_region.filter(indicateur=ind).aggregate(
                    Sum('valeur_realisee')
                )['valeur_realisee__sum'] or 0
                pct = (total_realise / ind.cible_finale * 100)
                total_pct_region += pct
                count_ind_region += 1
        
        pct_region = (total_pct_region / count_ind_region) if count_ind_region > 0 else 0
        
        region_data.append([region, str(nb_real), f'{pct_region:.1f}%'])
    
    region_table = Table(region_data, colWidths=[8*cm, 6*cm, 6*cm])
    region_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    story.append(region_table)
    story.append(PageBreak())
    
    # === SECTION 3: TOP 10 INDICATEURS ===
    story.append(Paragraph("3. TOP 10 INDICATEURS (PAR PERFORMANCE)", heading_style))
    
    indicateurs_performance = []
    
    for ind in indicateurs:
        if ind.cible_finale and ind.cible_finale > 0:
            total_realise = realisations.filter(indicateur=ind).aggregate(
                Sum('valeur_realisee')
            )['valeur_realisee__sum'] or 0
            pct = (total_realise / ind.cible_finale * 100)
            indicateurs_performance.append({
                'code': ind.code,
                'libelle': ind.libelle[:60],
                'pct': pct
            })
    
    # Trier par performance
    indicateurs_performance.sort(key=lambda x: x['pct'], reverse=True)
    top_10 = indicateurs_performance[:10]
    
    top_data = [['Code', 'Indicateur', 'Performance (%)']]
    for item in top_10:
        top_data.append([item['code'], item['libelle'], f"{item['pct']:.1f}%"])
    
    top_table = Table(top_data, colWidths=[4*cm, 12*cm, 4*cm])
    top_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('ALIGN', (2, 0), (2, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    story.append(top_table)
    story.append(Spacer(1, 1*cm))
    
    # === SECTION 4: ALERTES QUALITÉ ===
    story.append(Paragraph("4. ALERTES QUALITÉ NON RÉSOLUES", heading_style))
    
    alertes = AlerteQualite.objects.filter(resolue=False).select_related(
        'realisation__indicateur'
    ).order_by('-severite', '-date_detection')[:15]
    
    if alertes:
        alerte_data = [['Région', 'Indicateur', 'Type', 'Sévérité']]
        
        for alerte in alertes:
            alerte_data.append([
                alerte.realisation.region,
                alerte.realisation.indicateur.code,
                alerte.get_type_alerte_display(),
                alerte.get_severite_display()
            ])
        
        alerte_table = Table(alerte_data, colWidths=[5*cm, 5*cm, 5*cm, 5*cm])
        alerte_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        story.append(alerte_table)
    else:
        story.append(Paragraph("Aucune alerte non résolue.", styles['Normal']))
    
    # Pied de page
    story.append(Spacer(1, 2*cm))
    footer = Paragraph(
        f"<i>Rapport généré automatiquement le {datetime.now().strftime('%d/%m/%Y à %H:%M')}</i><br/>"
        "<i>ProSMAT - Projet de Sécurité Alimentaire et Nutritionnelle - GAFSP/FIDA</i>",
        ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, alignment=TA_CENTER)
    )
    story.append(footer)
    
    # Construire le PDF
    doc.build(story)
    
    # Préparer la réponse HTTP
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    filename = f'ProSMAT_Rapport_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response
